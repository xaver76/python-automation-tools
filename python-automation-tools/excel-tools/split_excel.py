# split_excel.py
# Divide un archivo Excel grande en varios archivos más pequeños.
# Requiere: pip install openpyxl

from openpyxl import load_workbook, Workbook
from pathlib import Path
import argparse


def split_excel(input_file, output_dir, rows_per_file=50000, sheet_name=None):
    input_path = Path(input_file)
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(input_path, read_only=True, data_only=True)

    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    rows = ws.iter_rows(values_only=True)

    try:
        header = next(rows)
    except StopIteration:
        print("El archivo está vacío.")
        return

    file_count = 1
    row_count = 0

    new_wb = Workbook(write_only=True)
    new_ws = new_wb.create_sheet(title="Datos")
    new_ws.append(header)

    for row in rows:
        if row_count >= rows_per_file:
            output_file = output_path / f"{input_path.stem}_parte_{file_count}.xlsx"
            new_wb.save(output_file)
            print(f"Generado: {output_file}")

            file_count += 1
            row_count = 0

            new_wb = Workbook(write_only=True)
            new_ws = new_wb.create_sheet(title="Datos")
            new_ws.append(header)

        new_ws.append(row)
        row_count += 1

    if row_count > 0:
        output_file = output_path / f"{input_path.stem}_parte_{file_count}.xlsx"
        new_wb.save(output_file)
        print(f"Generado: {output_file}")

    wb.close()
    print("Proceso finalizado.")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Divide un archivo Excel grande en varios archivos más pequeños."
    )

    parser.add_argument("input_file", help="Ruta del archivo Excel original")
    parser.add_argument("output_dir", help="Carpeta donde se guardarán los archivos divididos")
    parser.add_argument(
        "--rows",
        type=int,
        default=50000,
        help="Cantidad de filas por archivo. Default: 50000"
    )
    parser.add_argument(
        "--sheet",
        default=None,
        help="Nombre de la hoja a procesar. Si no se indica, usa la primera hoja activa."
    )

    args = parser.parse_args()

    split_excel(
        input_file=args.input_file,
        output_dir=args.output_dir,
        rows_per_file=args.rows,
        sheet_name=args.sheet
    )